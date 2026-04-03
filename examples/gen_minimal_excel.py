from __future__ import annotations

import asyncio
import os
import warnings
from pathlib import Path
from typing import Any

from deli_api import CaseHit, init_case_client
from dotenv import load_dotenv
from langchain_openai import ChatOpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from pydantic import BaseModel, Field


# =========================
# Debug Config (edit here)
# =========================
QUERY = "上班途中车祸工伤案例"
TARGET_CASE_COUNT = 50
PAGE_SIZE = 5  # Deli API 每页最多 5 条
MAX_CONCURRENCY = 5
OUTPUT_XLSX = Path("outputs/minimal_case_table.xlsx")
DOTENV_PATH = Path(r"f:\code_lib\YiZhi\.env")


HEADERS = [
    "审理法院",
    "案号",
    "法院认为（精简后）",
    "裁判结果",
    "主要原因",
    "关键裁判结论（权利类/金额类/行为类）",
]


class KeyConclusions(BaseModel):
    权利类: str = Field(default="未明确", description="权利义务方向的结论")
    金额类: str = Field(default="未明确", description="金额方向的结论")
    行为类: str = Field(default="未明确", description="行为义务方向的结论")


class CaseExtractResult(BaseModel):
    法院认为_精简后: str = Field(
        alias="法院认为（精简后）",
        description="法院认为部分的精简文本",
    )
    裁判结果: str = Field(description="判决主文/裁判结果")
    主要原因: str = Field(description="主要裁判原因，1-3句")
    关键裁判结论: KeyConclusions = Field(description="按三类输出的关键裁判结论")


def load_env() -> None:
    dotenv_path = DOTENV_PATH if DOTENV_PATH.exists() else None
    load_dotenv(dotenv_path=dotenv_path, override=False)


def init_llm() -> ChatOpenAI:
    api_key = os.getenv("LLM_API_KEY")
    if not api_key:
        raise RuntimeError("Missing LLM_API_KEY in environment.")

    return ChatOpenAI(
        model=os.getenv("LLM_MODEL"),
        api_key=api_key,
        temperature=0.1,
        base_url=os.getenv("LLM_BASE_URL"),
    )


def build_llm_extractor(llm: ChatOpenAI):
    return llm.with_structured_output(CaseExtractResult)


def _suppress_known_pydantic_warning() -> None:
    warnings.filterwarnings(
        "ignore",
        message=r"^Pydantic serializer warnings:",
        category=UserWarning,
        module=r"pydantic\.main",
    )


async def extract_with_llm(
    extractor: Any,
    case_title: str,
    raw_case_text: str,
) -> tuple[str, str, str, str]:
    prompt = f"""
你是法律助理。请严格基于给定内容提取字段，不要编造。
如果信息不足，请写“未明确”。

案件标题: {case_title}
裁判文书原文:
{raw_case_text}
    """.strip()

    _suppress_known_pydantic_warning()
    result = await extractor.ainvoke(prompt)
    court_reasoning_short = result.法院认为_精简后.strip() or "未明确"
    judgment_result = result.裁判结果.strip() or "未明确"
    major_reason = result.主要原因.strip() or "未明确"

    key_obj = result.关键裁判结论
    key_text = (
        f"权利类:{key_obj.权利类.strip()}；"
        f"金额类:{key_obj.金额类.strip()}；"
        f"行为类:{key_obj.行为类.strip()}"
    )
    return court_reasoning_short, judgment_result, major_reason, key_text


async def process_hit(
    extractor: Any,
    hit: CaseHit,
    semaphore: asyncio.Semaphore,
) -> list[str]:
    async with semaphore:
        court = str(hit.court_name or "")
        case_no = str(hit.case_no or hit.citation or "")
        raw_case_text = str(hit.content or "")

        (
            court_reasoning_short,
            judgment_result,
            major_reason,
            key_conclusion,
        ) = await extract_with_llm(
            extractor=extractor,
            case_title=hit.title,
            raw_case_text=raw_case_text,
        )

        return [
            court,
            case_no,
            court_reasoning_short,
            judgment_result,
            major_reason,
            key_conclusion,
        ]


def _case_key(hit: CaseHit) -> str:
    return str(hit.source_id or hit.case_no or hit.citation or hit.title)


async def rewrite_query(llm: ChatOpenAI, query: str) -> str:
    prompt = f"""
你是法律检索助手。请将用户输入改写为更专业、规范的法律检索语句。
只输出一行改写后的检索语句，不要解释，不要编号，不要引号。

用户输入：{query}
    """.strip()
    msg = await llm.ainvoke(prompt)
    rewritten = str(msg.content).strip().splitlines()[0].strip() if msg.content else ""
    return rewritten or query


async def fetch_case_hits(case_client: Any, query: str, target_count: int) -> list[CaseHit]:
    if target_count <= 0:
        return []

    page_size = max(1, min(PAGE_SIZE, 5))
    page_no = 1
    hits: list[CaseHit] = []
    seen: set[str] = set()

    while len(hits) < target_count:
        page_hits = await asyncio.to_thread(
            case_client.search_cases,
            query,
            page_no=page_no,
            page_size=page_size,
        )
        if not page_hits:
            break

        for hit in page_hits:
            key = _case_key(hit)
            if key in seen:
                continue
            seen.add(key)
            hits.append(hit)
            if len(hits) >= target_count:
                break

        if len(page_hits) < page_size:
            break
        page_no += 1

    return hits[:target_count]


async def build_rows() -> list[list[str]]:
    with init_case_client() as case_client:
        llm = init_llm()
        extractor = build_llm_extractor(llm)

        rewritten_query = await rewrite_query(llm, QUERY)
        original_hits = await fetch_case_hits(case_client, QUERY, TARGET_CASE_COUNT)
        rewritten_hits = (
            await fetch_case_hits(case_client, rewritten_query, TARGET_CASE_COUNT)
            if rewritten_query != QUERY
            else []
        )

        merged_hits: list[CaseHit] = []
        seen: set[str] = set()
        for i in range(max(len(original_hits), len(rewritten_hits))):
            if i < len(original_hits):
                key = _case_key(original_hits[i])
                if key not in seen:
                    seen.add(key)
                    merged_hits.append(original_hits[i])
            if len(merged_hits) >= TARGET_CASE_COUNT:
                break
            if i < len(rewritten_hits):
                key = _case_key(rewritten_hits[i])
                if key not in seen:
                    seen.add(key)
                    merged_hits.append(rewritten_hits[i])
            if len(merged_hits) >= TARGET_CASE_COUNT:
                break

        semaphore = asyncio.Semaphore(MAX_CONCURRENCY)
        tasks = [process_hit(extractor, hit, semaphore) for hit in merged_hits]
        return await asyncio.gather(*tasks)


def _save_excel_sync(rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "类案检索结果"

    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [16, 24, 40, 36, 36, 52]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)


async def save_excel(rows: list[list[str]]) -> None:
    await asyncio.to_thread(_save_excel_sync, rows)


async def main() -> None:
    load_env()
    rows = await build_rows()
    await save_excel(rows)
    print(f"Generated: {OUTPUT_XLSX.resolve()}")
    print(f"Rows: {len(rows)}")
    print("Columns:", " | ".join(HEADERS))


if __name__ == "__main__":
    asyncio.run(main())
