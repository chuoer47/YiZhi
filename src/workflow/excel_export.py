from __future__ import annotations

import asyncio
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

if __package__:
    from .excel_models import MinimalCaseRow
    from .workflow_config import HEADERS
else:
    from excel_models import MinimalCaseRow
    from workflow_config import HEADERS


def save_excel_sync(rows: list[MinimalCaseRow]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "类案检索结果"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row.to_excel_row())

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [16, 24, 20, 16, 14, 40, 34, 28, 22, 22, 22, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path = Path("outputs") / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


async def save_excel(rows: list[MinimalCaseRow]) -> Path:
    return await asyncio.to_thread(save_excel_sync, rows)
