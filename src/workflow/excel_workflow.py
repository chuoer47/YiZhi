from __future__ import annotations

import asyncio
import html
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

# Allow direct run: python src/workflow/excel_workflow.py
SRC_ROOT = Path(__file__).resolve().parents[1]
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from deli_api import CaseHit, DeliLegalClient, LawHit, init_case_client
from dotenv import load_dotenv
from langchain_core.messages import HumanMessage, SystemMessage
from langchain_openai import ChatOpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from pydantic import BaseModel, ConfigDict, Field
if __package__:
    from .prompt_loader import render_prompt
else:
    from prompt_loader import render_prompt


QUERY = "上班途中车祸工伤案例"
TARGET_CASE_COUNT = 50
REWRITE_QUERY_COUNT = 3
PAGE_SIZE = 5
MAX_CONCURRENCY = 5
DOTENV_PATH = Path(".env")

HEADERS = [
    "审理法院",
    "案号",
    "案由",
    "法院审理程序",
    "法院层级",
    "法院认为（精简后）",
    "裁判结果",
    "主要原因",
    "关键裁判结论-权利类",
    "关键裁判结论-金额类",
    "关键裁判结论-行为类",
    "法律依据（法律名称+条文）",
]


class KeyConclusions(BaseModel):
    权利类: str = Field(..., min_length=1)
    金额类: str = Field(..., min_length=1)
    行为类: str = Field(..., min_length=1)


class CaseExtractResult(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    案由: str = Field(..., min_length=1)
    法院审理程序: str = Field(..., min_length=1)
    法院层级: str = Field(..., min_length=1)
    法院认为_精简后: str = Field(..., alias="法院认为（精简后）", min_length=1)
    裁判结果: str = Field(..., min_length=1)
    主要原因: str = Field(..., min_length=1)
    关键裁判结论: KeyConclusions = Field(...)


class QueryRewriteResult(BaseModel):
    改写查询: list[str] = Field(..., min_length=REWRITE_QUERY_COUNT, max_length=REWRITE_QUERY_COUNT)


class LawQueryRewriteResult(BaseModel):
    法律检索查询: list[str] = Field(..., min_length=REWRITE_QUERY_COUNT, max_length=REWRITE_QUERY_COUNT)


class MinimalCaseRow(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    审理法院: str = Field(..., min_length=1)
    案号: str = Field(..., min_length=1)
    案由: str = Field(..., min_length=1)
    法院审理程序: str = Field(..., min_length=1)
    法院层级: str = Field(..., min_length=1)
    法院认为_精简后: str = Field(..., alias="法院认为（精简后）", min_length=1)
    裁判结果: str = Field(..., min_length=1)
    主要原因: str = Field(..., min_length=1)
    关键裁判结论_权利类: str = Field(..., min_length=1)
    关键裁判结论_金额类: str = Field(..., min_length=1)
    关键裁判结论_行为类: str = Field(..., min_length=1)
    法律依据: str = Field(..., alias="法律依据（法律名称+条文）", min_length=1)

    def to_excel_row(self) -> list[str]:
        return [
            self.审理法院,
            self.案号,
            self.案由,
            self.法院审理程序,
            self.法院层级,
            self.法院认为_精简后,
            self.裁判结果,
            self.主要原因,
            self.关键裁判结论_权利类,
            self.关键裁判结论_金额类,
            self.关键裁判结论_行为类,
            self.法律依据,
        ]


def load_env() -> None:
    load_dotenv(dotenv_path=DOTENV_PATH, override=False)


def init_llm() -> ChatOpenAI:
    api_key = os.getenv("LLM_API_KEY")
    model = os.getenv("LLM_MODEL")
    base_url = os.getenv("LLM_BASE_URL")
    if not api_key or not model or not base_url:
        missing = [
            key
            for key, value in (
                ("LLM_API_KEY", api_key),
                ("LLM_MODEL", model),
                ("LLM_BASE_URL", base_url),
            )
            if not value
        ]
        raise RuntimeError(f"Missing required LLM env vars: {', '.join(missing)}")
    return ChatOpenAI(model=model, api_key=api_key, temperature=0.1, base_url=base_url)


def build_case_dedup_key(hit: CaseHit) -> str:
    case_no = (hit.case_no or "").replace("（", "(").replace("）", ")").replace(" ", "")
    if not case_no:
        raise ValueError(f"Case missing case_no: {hit.title}")
    return case_no


def build_law_basis(hits: list[LawHit]) -> str:
    parts: list[str] = []
    for hit in hits:
        law_name = re.sub(r"<[^>]+>", "", html.unescape(hit.law_name or "")).strip()
        article_no = re.sub(r"<[^>]+>", "", html.unescape(hit.article_no or "")).strip()
        if law_name and article_no:
            parts.append(f"{law_name}{article_no}")
        if len(parts) >= 2:
            break
    return "；".join(parts)


async def rewrite_queries(rewriter: Any, query: str) -> list[str]:
    if not query.strip():
        raise ValueError("QUERY must not be empty.")
    schema_text = json.dumps(QueryRewriteResult.model_json_schema(), ensure_ascii=False, indent=2)
    result: QueryRewriteResult = await rewriter.ainvoke(
        [
            SystemMessage(content=render_prompt("excel/rewrite_queries_system.txt")),
            HumanMessage(
                content=render_prompt(
                    "excel/rewrite_queries_human.txt",
                    query=query,
                    rewrite_query_count=REWRITE_QUERY_COUNT,
                    schema_text=schema_text,
                )
            ),
        ]
    )
    rewritten = [item.strip() for item in result.改写查询]
    if len(set(rewritten)) != len(rewritten):
        raise ValueError("Rewritten queries contain duplicates.")
    if any(item == query for item in rewritten):
        raise ValueError("Rewritten queries contain the original query.")
    return rewritten


async def fetch_case_hits(case_client: Any, query: str, target_count: int) -> list[CaseHit]:
    if target_count <= 0:
        raise ValueError("TARGET_CASE_COUNT must be greater than 0.")
    if PAGE_SIZE < 1 or PAGE_SIZE > 5:
        raise ValueError("PAGE_SIZE must be between 1 and 5.")
    page_no = 1
    hits: list[CaseHit] = []
    seen: set[str] = set()
    while len(hits) < target_count:
        page_hits = await asyncio.to_thread(
            case_client.search_cases,
            query,
            page_no=page_no,
            page_size=PAGE_SIZE,
        )
        if not page_hits:
            break
        for hit in page_hits:
            key = build_case_dedup_key(hit)
            if key in seen:
                continue
            seen.add(key)
            hits.append(hit)
            if len(hits) >= target_count:
                break
        if len(page_hits) < PAGE_SIZE:
            break
        page_no += 1
    return hits


async def extract_case_fields(extractor: Any, hit: CaseHit) -> CaseExtractResult:
    if not hit.title.strip():
        raise ValueError("Case title is empty.")
    if not hit.content.strip():
        raise ValueError(f"Case content is empty: {hit.title}")
    schema_text = json.dumps(CaseExtractResult.model_json_schema(), ensure_ascii=False, indent=2)
    return await extractor.ainvoke(
        [
            SystemMessage(content=render_prompt("excel/extract_case_fields_system.txt")),
            HumanMessage(
                content=render_prompt(
                    "excel/extract_case_fields_human.txt",
                    case_title=hit.title,
                    court_name=hit.court_name or "",
                    case_no=hit.case_no or "",
                    cause=hit.cause or "",
                    level_of_trial=hit.level_of_trial or "",
                    case_type=hit.case_type or "",
                    case_content=hit.content,
                    schema_text=schema_text,
                )
            ),
        ]
    )


async def rewrite_law_queries(law_rewriter: Any, hit: CaseHit, extracted: CaseExtractResult) -> list[str]:
    schema_text = json.dumps(LawQueryRewriteResult.model_json_schema(), ensure_ascii=False, indent=2)
    result: LawQueryRewriteResult = await law_rewriter.ainvoke(
        [
            SystemMessage(content=render_prompt("excel/rewrite_law_queries_system.txt")),
            HumanMessage(
                content=render_prompt(
                    "excel/rewrite_law_queries_human.txt",
                    case_title=hit.title,
                    cause=extracted.案由,
                    major_reason=extracted.主要原因,
                    judgment_result=extracted.裁判结果,
                    schema_text=schema_text,
                )
            ),
        ]
    )
    rewritten = [item.strip() for item in result.法律检索查询 if item.strip()]
    base_queries = [
        f"{hit.title} {extracted.主要原因}".strip(),
        extracted.案由.strip(),
        (hit.cause or "").strip(),
        extracted.主要原因.strip(),
    ]
    merged: list[str] = []
    seen: set[str] = set()
    for q in [*rewritten, *base_queries]:
        if q and q not in seen:
            seen.add(q)
            merged.append(q)
    return merged


async def process_hit(
    extractor: Any,
    law_rewriter: Any,
    law_client: DeliLegalClient,
    hit: CaseHit,
    semaphore: asyncio.Semaphore,
) -> MinimalCaseRow:
    async with semaphore:
        if not hit.court_name or not hit.case_no:
            raise ValueError(f"Case metadata missing court_name/case_no: {hit.title}")
        extracted = await extract_case_fields(extractor, hit)
        law_queries = await rewrite_law_queries(law_rewriter, hit, extracted)
        law_basis = ""
        for query in law_queries:
            q = query.strip()
            if not q:
                continue
            law_hits = await asyncio.to_thread(
                law_client.search_laws,
                q,
                page_size=3,
                field_name="semantic",
            )
            law_basis = build_law_basis(law_hits)
            if law_basis:
                break
        if not law_basis:
            law_basis = "未检索到明确法律条文"
        return MinimalCaseRow(
            审理法院=hit.court_name.strip(),
            案号=hit.case_no.strip(),
            案由=extracted.案由.strip(),
            法院审理程序=extracted.法院审理程序.strip(),
            法院层级=extracted.法院层级.strip(),
            法院认为_精简后=extracted.法院认为_精简后.strip(),
            裁判结果=extracted.裁判结果.strip(),
            主要原因=extracted.主要原因.strip(),
            关键裁判结论_权利类=extracted.关键裁判结论.权利类.strip(),
            关键裁判结论_金额类=extracted.关键裁判结论.金额类.strip(),
            关键裁判结论_行为类=extracted.关键裁判结论.行为类.strip(),
            法律依据=law_basis.strip(),
        )


async def build_rows(
    query: str = QUERY,
    target_case_count: int = TARGET_CASE_COUNT,
) -> tuple[list[MinimalCaseRow], list[CaseHit]]:
    if not query.strip():
        raise ValueError("query must not be empty.")
    if target_case_count <= 0:
        raise ValueError("target_case_count must be greater than 0.")

    with init_case_client() as case_client, DeliLegalClient.from_env(DOTENV_PATH) as law_client:
        llm = init_llm()
        rewriter = llm.with_structured_output(QueryRewriteResult, method="function_calling")
        law_rewriter = llm.with_structured_output(LawQueryRewriteResult, method="function_calling")
        extractor = llm.with_structured_output(CaseExtractResult, method="function_calling")

        rewritten_queries = await rewrite_queries(rewriter, query)
        all_queries = [query, *rewritten_queries]

        merged_hits: list[CaseHit] = []
        seen: set[str] = set()
        for search_query in all_queries:
            hits = await fetch_case_hits(case_client, search_query, target_case_count)
            for hit in hits:
                key = build_case_dedup_key(hit)
                if key in seen:
                    continue
                seen.add(key)
                merged_hits.append(hit)
                if len(merged_hits) >= target_case_count:
                    break
            if len(merged_hits) >= target_case_count:
                break

        semaphore = asyncio.Semaphore(MAX_CONCURRENCY)
        tasks = [process_hit(extractor, law_rewriter, law_client, hit, semaphore) for hit in merged_hits]
        rows = await asyncio.gather(*tasks)
        return rows, merged_hits


def save_excel_sync(rows: list[MinimalCaseRow]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "类案检索结果"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row.to_excel_row())

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [16, 24, 20, 16, 14, 40, 34, 28, 22, 22, 22, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path = Path("outputs") / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


async def save_excel(rows: list[MinimalCaseRow]) -> Path:
    return await asyncio.to_thread(save_excel_sync, rows)


async def run_excel_workflow(
    query: str = QUERY,
    target_case_count: int = TARGET_CASE_COUNT,
) -> tuple[Path, list[MinimalCaseRow], list[CaseHit]]:
    load_env()
    rows, case_hits = await build_rows(query=query, target_case_count=target_case_count)
    excel_path = await save_excel(rows)
    return excel_path, rows, case_hits


async def main() -> None:
    excel_path, rows, _ = await run_excel_workflow()
    print(f"Generated Excel: {excel_path.resolve()}")
    print(f"Rows: {len(rows)}")
    print("Columns:", " | ".join(HEADERS))


if __name__ == "__main__":
    asyncio.run(main())
